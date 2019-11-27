#!/bin/python3
from openpyxl.styles import NamedStyle
import datetime
import math
from openpyxl import Workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet(title="timepoints")
# datetime has 19 characters
ws.column_dimensions['A'].width = 19
ws.column_dimensions['B'].width = 19


start = (datetime.datetime.now() - datetime.timedelta(days=10)) - datetime.timedelta(seconds=320)

start = start.replace(microsecond=0, second=0, minute=0, hour=0)

channels = [1, 2, 3, 4, 5, 6, 7, 8]
maxValue = 100

end = start + datetime.timedelta(days=1)
interval = datetime.timedelta(minutes=1)
# daylength = datetime.timedelta(hours=3)

ws.append(["datetime", "datetime-sim", "humidity", "temperature", "light1", "light2"])

x = 0


def is_night(dt):
    return not (7 <= dt.hour < 23)


a = 0
while start < end:
    a += 1
    dt = start
    start += datetime.timedelta(seconds=10)
    if is_night(dt):
        # start += datetime.timedelta(minutes=10)
        light1 = light2 = 0
        temp = 20 if start.hour % 2 == 0 else 25
    else:
        temp = 28 if start.hour % 2 == 0 else 24
        # if a % 2 == 0:
        #     start += datetime.timedelta(minutes=10)
        #     light1 = light2 = 2

        # else:
        #     light1 = 5
        #     light2 = 3
        #     start += datetime.timedelta(minutes=2)

    ws.append([
        dt,
        dt,
        55,
        temp,
        light1,
        light2
    ])
    if start >= datetime.datetime.now() - datetime.timedelta(days=10):
        break

# set datetime format to a more iso8601
# currently comes out 'YYYY-MM-DD H:MM:MM' which is wrong.
# this doesnt work to do it,
# date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:MM')
# for cell in ws['A'][1:]:
#     cell.style = date_style
# for cell in ws['B'][1:]:
#     cell.style = date_style


wb.save(filename="timepoints.xlsx")
